import os, sys
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas
from pandas.plotting import parallel_coordinates
import numpy as np
from matplotlib import ticker

# benalexkeen parallel coordinates
FOLDER_NAME = 0
IMAGE = 1
ATTACK_TYPE = 2
SUCCESS = 3
SUCCESS_TYPE = 4

WIDTH_MIN = 0.1 # width of shadow from perspective of vehicle (m)
WIDTH_MAX = 3.6

LENGTH_MIN = 0.5 # length of shadow from perspective of vehicle (m)
LENGTH_MAX = 15

BETA_MIN = 0 # rotation of negative shadow relative to lane marker (degrees)
BETA_MAX = 90

TRANSPARENCY_MIN = 90 # transparency of positive shadow
TRANSPARENCY_MAX = 90

BLUR_MIN = 0 # degree of blur (softness) of positive shadow
BLUR_MAX = 0

DISTANCE_MIN = 0.1 # distance of negative shadow relative to lane marker (m)
DISTANCE_MAX = 3.6

DIV_NUM = 1000 # number of rows per csv

def make_list(min: float, max: float, step: float) -> list:
    """ Creates a list based on input params, rounds to second decimal place
    """
    return [round(x, 2) for x in np.arange(min, max, step)]



def build_results(workbook: Workbook, worksheet_name: str, attr_csv: pandas.DataFrame) -> list:
    """ Builds a dict for a given worksheet containing the results and subsequent
    shadow attrs

    Args:
        workbook: loaded workbook to analyze
        worksheet_name: name of worksheet to grab values from
        attr_csv: csv loaded as pandas DF to get values from

    """
    
    ws = workbook[worksheet_name]
    results = {}

    results['mid.png'] = {}
    for row in ws.iter_rows(min_row=2, max_col=5, max_row=ws.max_row):
        if row[IMAGE].value not in results:
            results[row[IMAGE].value] = {}

        row_results = {'folder_name': row[FOLDER_NAME].value,
                       'attack_type': int(row[ATTACK_TYPE].value),
                       'success': int(row[SUCCESS].value),
                       'success_type': row[SUCCESS_TYPE].value,
                       'attr_vals': attr_csv.iloc[int(row[ATTACK_TYPE].value)]} 

        results[row[IMAGE].value].update({row[ATTACK_TYPE].value: row_results})

    return results


def make_param_dict(param_list: list) -> dict:
    """
    """
    param = {}
    for val in param_list:
        param.update({val: {'success': 0, 'failure': 0}})

    return param


def extract_data(results_dict):
    """
    """
    extract = {}
    for img in results_dict:
        if img not in extract:
            extract[img] = []
        for result in results_dict[img]:
            df = results_dict[img][result]['attr_vals']
            data_tup = (df['width'], df['length'], df['beta'], df['distance'], results_dict[img][result]['success'])
            extract[img].append(data_tup)

    return extract
            
def update_points(wb, worksheet_name, csv_df, points_dict):
    results = build_results(wb, worksheet_name, csv_df)
    data = extract_data(results)

    for img in points_dict:
        points_dict[img].extend(data[img])

def determine_stats(df: pandas.DataFrame):
    """
    """
    x = df['width']
    y = df['length']
    z = df['beta']
    c = df['success']

    fig, ax1 = plt.subplots()
    ax1.set_xlabel('success')
    ax1.set_ylabel('width', color='tab:red')
    ax1.bar(c, x, color='tab:red')
    ax1.tick_params(axis='y', labelcolor='tab:red')

    # ax2 = ax1.twinx()
    # color = 'tab:blue'
    # ax2.set_ylabel('length')
    # ax2.bar(c, y, color=color)
    # ax2.tick_params(axis='y', labelcolor=color)
    # fig.tight_layout()

    plt.show()

def plotting(points):
    df = pandas.DataFrame(points['mid.png'], columns=['width', 'length', 'beta', 'distance', 'success'])
    
    fig = plt.figure()
    ax = plt.axes(projection='3d')
    ax.scatter3D(xs=df['width'], ys=df['success'], zs=df['length'])
    ax.set_xlabel('Width')
    ax.set_ylabel('Success')
    ax.set_zlabel('Theta')

    plt.title('test')
    plt.show()

def set_ticks_for_axes(dim, ax, ticks, min_max_range, cols, df):
    min_val, max_val, val_range = min_max_range[cols[dim]]
    step = val_range / float(ticks - 1)
    tick_labels = [round(min_val + step * i, 2) for i in range(ticks)]
    norm_min = df[cols[dim]].min()
    norm_range = np.ptp(df[cols[dim]])
    norm_step = norm_range / float(ticks-1)
    ticks = [round(norm_min + norm_step * i, 2) for i in range(ticks)]
    ax.yaxis.set_ticks(ticks)
    ax.set_yticklabels(tick_labels)

def parallel_coords(df):
    cols = ['width', 'length', 'distance', 'beta', 'success']
    df['success'] = df['success'].astype('category')
    x = [i for i, _ in enumerate(cols)]


    colors = ['#2e8ad8', '#cd3785']
    colors = {df['success'].cat.categories[i]: colors[i] for i, _ in enumerate(df['success'].cat.categories)}
    fig, axes = plt.subplots(1, len(x) - 1, sharey=False, figsize=(10,5))

    min_max_range = {}
    for col in cols[:-1]:
        min_max_range[col] = [df[col].min(), df[col].max(), np.ptp(df[col])]
        df[col] = np.true_divide(df[col] - df[col].min(), np.ptp(df[col]))

    min_max_range['success'] = [0, 1, 1]
    print(min_max_range)

    for i, ax in enumerate(axes):
        for idx in df.index:
            success_cat = df.loc[idx,'success']
            ax.plot(x, df.loc[idx, cols], colors[success_cat])
        
        ax.set_xlim([x[i], x[i+1]])
    
    for dim, ax in enumerate(axes):
        ax.xaxis.set_major_locator(ticker.FixedLocator([dim]))
        print(f'{cols[dim]} {dim} {df[cols[dim]].dtype}')

        set_ticks_for_axes(dim, ax, 6, min_max_range, cols, df)
        ax.set_xticklabels([cols[dim]])

    ax = plt.twinx(axes[-1])
    dim = len(axes)
    ax.xaxis.set_major_locator(ticker.FixedLocator([x[-2], x[-1]]))

    ax.yaxis.set_ticks([0,1])
    ax.set_yticklabels([0,1])
    ax.set_xticklabels([cols[-2], cols[-1]])

    plt.subplots_adjust(wspace=0)
    plt.legend([plt.Line2D((0,1), (0,0), color=colors[cat]) for cat in df['success'].cat.categories],
               df['success'].cat.categories,
               bbox_to_anchor=(1.2, 1), loc=2, borderaxespad=0.)
    plt.show() 

    


def analyze(config: dict):
    """
    """
    d = pandas.read_csv('./Shadow_Attack/masterscript/data/panda/clrernetback_edge.png.csv')
    parallel_coords(d)
    # clrernet_points = {'mid.png': [], 'back_edge.png': [], 'mid_lane.png': [], 'front_edge.png': []}
    # hybridnets_points = {'mid.png': [], 'back_edge.png': [], 'mid_lane.png': [], 'front_edge.png': []}
    # twinlitenet_points = {'mid.png': [], 'back_edge.png': [], 'mid_lane.png': [], 'front_edge.png': []}

    # results_dir = config['results_folder']
    # for subdir, _, files in os.walk(results_dir):
    #     if files:
    #         csv_file = [x for x in files if '.csv' in x and 'lock' not in x][0]
    #         xlsx_file = [x for x in files if '.xlsx' in x and 'lock' not in x][0]
    #         num = csv_file.split('.')[0].split('shadow_attrs')[1]

    #         if f'Results{num}' not in xlsx_file:
    #             print(f'Error: File system is wrong: {csv_file} {xlsx_file}')
    #             sys.exit(0)

    #         wb = load_workbook(os.path.join(subdir, xlsx_file))
    #         csv_df = pandas.read_csv(os.path.join(subdir, csv_file))

    #         update_points(wb, 'crlernet_results', csv_df, clrernet_points)
    #         update_points(wb, 'hybridnets', csv_df, hybridnets_points)
    #         update_points(wb, 'twinlitenet', csv_df, twinlitenet_points)

    # temp = os.path.join(config['data_dir'], 'panda')
    # os.makedirs(temp,  exist_ok=True)

    # for img in clrernet_points:
    #     df = pandas.DataFrame(clrernet_points[img], columns=['width', 'length', 'beta', 'distance', 'success'])
    #     df.to_csv(os.path.join(temp,f'clrernet{img}.csv'), index=False)

    # for img in hybridnets_points:
    #     df = pandas.DataFrame(hybridnets_points[img], columns=['width', 'length', 'beta', 'distance', 'success'])
    #     df.to_csv(os.path.join(temp,f'hybridnets{img}.csv'), index=False)

    # for img in twinlitenet_points:
    #        df = pandas.DataFrame(twinlitenet_points[img], columns=['width', 'length', 'beta', 'distance', 'success'])
    #        df.to_csv(os.path.join(temp,f'twinlitenet{img}.csv'), index=False)

# analyze('')