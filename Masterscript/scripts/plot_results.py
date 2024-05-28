import matplotlib.pyplot as plt
import numpy as np
import pandas
import os,sys
import math
from openpyxl import Workbook, load_workbook
from pandas.plotting import parallel_coordinates
from matplotlib import ticker
import plotly.express as px
import plotly.graph_objects as go



WORK_DIR = './Shadow_Attack/masterscript/data/panda/'
CLRNET_BACK = os.path.join(WORK_DIR, 'clrernetback_edge.csv')
CLRNET_FRONT = os.path.join(WORK_DIR, 'clrernetfront_edge.csv')
CLRNET_MID = os.path.join(WORK_DIR, 'clrernetmid.csv')
CLRNET_MID_LANE = os.path.join(WORK_DIR, 'clrernetmid_lane.csv')

HYBN_BACK = os.path.join(WORK_DIR, 'hybridnetsback_edge.csv')
HYBN_FRONT = os.path.join(WORK_DIR, 'hybridnetsfront_edge.csv')
HYBN_MID = os.path.join(WORK_DIR, 'hybridnetsmid.csv')
HYBN_MID_LANE = os.path.join(WORK_DIR, 'hybridnetsmid_lane.csv')

TWIN_BACK = os.path.join(WORK_DIR, 'twinlitenetback_edge.csv')
TWIN_FRONT = os.path.join(WORK_DIR, 'twinlitenetfront_edge.csv')
TWIN_MID = os.path.join(WORK_DIR, 'twinlitenetmid.csv')
TWIN_MID_LANE = os.path.join(WORK_DIR, 'twinlitenetmid_lane.csv')

FOLDER_NAME = 0
IMAGE = 1
ATTACK_TYPE = 2
SUCCESS = 3
SUCCESS_TYPE = 4

MAX_X = 4


def find_x2(row):
    """
    """
    distance = row['distance']
    width = row['width']
    length = row['length']
    beta = math.radians(90 - row['beta'])

    return distance + width * math.sin(beta) + length * math.cos(beta)


def clean_data():
    for filename in os.listdir(WORK_DIR):
        if filename.endswith('.csv'):
            df = pandas.read_csv(os.path.join(WORK_DIR, filename))
            print(f'Processing {filename} current row count: {df.shape[0]}')
            df['ns_width'] = df.apply(find_x2, axis=1)
            df = df.drop(df[(df['ns_width'] > MAX_X)].index)
            print(f'Post processing: {df.shape[0]}')

            avg_dist = df.groupby('distance').mean()
            avg_width = df.groupby('width').mean()

            fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2)

            ax1.plot(avg_dist.index, avg_dist['success'])
            ax1.set_title('Distance vs Mean success')
            ax1.set_xlabel('Distance')
            ax1.set_ylabel('Mean Success')

            ax2.plot(avg_dist.index, avg_dist['ns_width'])
            ax2.set_title('Distance vs Mean Negative Shadow X_2')
            ax2.set_xlabel('Distance (m)')
            ax2.set_ylabel('Mean Negative Shadow_x2 (m)')

            ax3.plot(avg_width.index, avg_width['success'])
            ax3.set_title('Width vs Mean Success')
            ax3.set_xlabel('Width (m)')
            ax3.set_ylabel('Mean Success')

            ax4.plot(avg_width.index, avg_width['ns_width'])
            ax4.set_title('Width vs Mean Negative Shadow X_2')
            ax4.set_xlabel('Width (m)')
            ax4.set_ylabel('Mean Negative Shadow_x2 (m)')

            fig.suptitle(f"{filename.split('.csv')[0]}, total samples: {df.shape[0]}\n"
                         f"Min X_2 {round(df['ns_width'].min(), 2)}m Max x_2 {round(df['ns_width'].max(), 2)}m")

            plt.show()



def plot_values():
    back = pandas.read_csv(CLRNET_BACK)
    avg_dist = back.groupby(['distance']).mean()
    avg_width = back.groupby(['width']).mean()
    avg_length = back.groupby(['length']).mean()
    avg_beta = back.groupby(['beta']).mean()

    fig, ax = plt.subplots()

    ax = avg_dist.plot(ax=ax, y='success', label='distance success')
    ax = avg_width.plot(ax=ax, y='success', label='width success')

    plt.legend()
    plt.show()

def build_results(workbook: Workbook, worksheet_name: str, attr_csv: pandas.DataFrame) -> list:
    """ Builds a dict for a given worksheet containing the results and subsequent
    shadow attrs

    Args:
        workbook: loaded workbook to analyze
        worksheet_name: name of worksheet to grab values from
        attr_csv: csv loaded as pandas DF to get values from

    """
    
    ws = workbook[worksheet_name]
    results = {}

    results['mid.png'] = {}
    for row in ws.iter_rows(min_row=2, max_col=5, max_row=ws.max_row):
        if row[IMAGE].value not in results:
            results[row[IMAGE].value] = {}

        row_results = {'folder_name': row[FOLDER_NAME].value,
                       'attack_type': int(row[ATTACK_TYPE].value),
                       'success': int(row[SUCCESS].value),
                       'success_type': row[SUCCESS_TYPE].value,
                       'attr_vals': attr_csv.iloc[int(row[ATTACK_TYPE].value)]} 

        results[row[IMAGE].value].update({row[ATTACK_TYPE].value: row_results})

    return results

def extract_data(results_dict):
    """
    """
    extract = {}
    for img in results_dict:
        if img not in extract:
            extract[img] = []
        for result in results_dict[img]:
            df = results_dict[img][result]['attr_vals']
            data_tup = (df['width'], df['length'], df['beta'], df['distance'], results_dict[img][result]['success'])
            extract[img].append(data_tup)

    return extract
            
def update_points(wb, worksheet_name, csv_df, points_dict):
    results = build_results(wb, worksheet_name, csv_df)
    data = extract_data(results)

    for img in points_dict:
        points_dict[img].extend(data[img])

def results_to_csv(results_dir, data_dir):
    clrernet_points = {'mid.png': [], 'back_edge.png': [], 'mid_lane.png': [], 'front_edge.png': []}
    hybridnets_points = {'mid.png': [], 'back_edge.png': [], 'mid_lane.png': [], 'front_edge.png': []}
    twinlitenet_points = {'mid.png': [], 'back_edge.png': [], 'mid_lane.png': [], 'front_edge.png': []}

    for subdir, _, files in os.walk(results_dir):
        if files:
            csv_file = [x for x in files if '.csv' in x and 'lock' not in x][0]
            xlsx_file = min([x for x in files if '.xlsx' in x and 'lock' not in x])
            num = csv_file.split('.')[0].split('shadow_attrs')[1]

            if f'Results{num}' not in xlsx_file:
                print(f'Error: File system is wrong: {csv_file} {xlsx_file}')
                sys.exit(0)

            wb = load_workbook(os.path.join(subdir, xlsx_file))
            csv_df = pandas.read_csv(os.path.join(subdir, csv_file))
            
            update_points(wb, 'clrernet_results', csv_df, clrernet_points)
            # update_points(wb, 'hybridnets', csv_df, hybridnets_points)
            # update_points(wb, 'twinlitenet', csv_df, twinlitenet_points)

    temp = os.path.join(data_dir, 'panda')
    os.makedirs(temp,  exist_ok=True)

    for img in clrernet_points:
        df = pandas.DataFrame(clrernet_points[img], columns=['width', 'length', 'beta', 'distance', 'success'])
        df.to_csv(os.path.join(temp,f'clrernet{img.split(".png")[0]}.csv'), index=False)

    # for img in hybridnets_points:
    #     df = pandas.DataFrame(hybridnets_points[img], columns=['width', 'length', 'beta', 'distance', 'success'])
    #     df.to_csv(os.path.join(temp,f'hybridnets{img.split(".png")[0]}.csv'), index=False)

    # for img in twinlitenet_points:
    #        df = pandas.DataFrame(twinlitenet_points[img], columns=['width', 'length', 'beta', 'distance', 'success'])
    #        df.to_csv(os.path.join(temp,f'twinlitenet{img.split(".png")[0]}.csv'), index=False)


def plot_distance():
    """ Plots the success of distance over the range of values

    """
    back = pandas.read_csv(CLRNET_BACK)
    avg_success = back.groupby(['distance']).mean()
    print(avg_success['success'])
    plt.figure()

    avg_success.plot(y='success')
    plt.show()

def combine_dfs():
    df_c_back = pandas.read_csv(CLRNET_BACK)
    df_c_front = pandas.read_csv(CLRNET_FRONT)
    df_c_mid = pandas.read_csv(CLRNET_MID)
    df_c_midlane = pandas.read_csv(CLRNET_MID_LANE)

    df_h_back = pandas.read_csv(HYBN_BACK)
    df_h_front = pandas.read_csv(HYBN_FRONT)
    df_h_mid = pandas.read_csv(HYBN_MID)
    df_h_midlane = pandas.read_csv(HYBN_MID_LANE)

    df_t_back = pandas.read_csv(TWIN_BACK)
    df_t_front = pandas.read_csv(TWIN_FRONT)
    df_t_mid = pandas.read_csv(TWIN_MID)
    df_t_midlane = pandas.read_csv(TWIN_MID_LANE)

    df_c = pandas.concat([df_c_back, df_c_front, df_c_mid, df_c_midlane])
    df_h = pandas.concat([df_h_back, df_h_front, df_h_mid, df_h_midlane])
    df_t = pandas.concat([df_t_back, df_t_front, df_t_mid, df_t_midlane])

    return df_c, df_h, df_t

def set_ticks_for_axes(dim, ax, ticks, min_max_range, cols, df):
    min_val, max_val, val_range = min_max_range[cols[dim]]
    step = val_range / float(ticks - 1)
    tick_labels = [round(min_val + step * i, 2) for i in range(ticks)]
    norm_min = df[cols[dim]].min()
    norm_range = np.ptp(df[cols[dim]])
    norm_step = norm_range / float(ticks-1)
    ticks = [round(norm_min + norm_step * i, 2) for i in range(ticks)]
    ax.yaxis.set_ticks(ticks)
    ax.set_yticklabels(tick_labels)

def plot_parallel(df):
    cols = ['width', 'length', 'distance', 'beta', 'success']
    df['success'] = df['success'].astype('category')
    x = [i for i, _ in enumerate(cols)]


    colors = ['#2e8ad8', '#cd3785']
    colors = {df['success'].cat.categories[i]: colors[i] for i, _ in enumerate(df['success'].cat.categories)}
    fig, axes = plt.subplots(1, len(x) - 1, sharey=False, figsize=(10,5))

    min_max_range = {}
    for col in cols[:-1]:
        min_max_range[col] = [df[col].min(), df[col].max(), np.ptp(df[col])]
        df[col] = np.true_divide(df[col] - df[col].min(), np.ptp(df[col]))

    min_max_range['success'] = [0, 1, 1]
    print(min_max_range)

    for i, ax in enumerate(axes):
        for idx in df.index:
            success_cat = df.loc[idx,'success']
            ax.plot(x, df.loc[idx, cols], colors[success_cat])
        
        ax.set_xlim([x[i], x[i+1]])
    
    for dim, ax in enumerate(axes):
        ax.xaxis.set_major_locator(ticker.FixedLocator([dim]))
        print(f'{cols[dim]} {dim} {df[cols[dim]].dtype}')

        set_ticks_for_axes(dim, ax, 6, min_max_range, cols, df)
        ax.set_xticklabels([cols[dim]])

    ax = plt.twinx(axes[-1])
    dim = len(axes)
    ax.xaxis.set_major_locator(ticker.FixedLocator([x[-2], x[-1]]))

    ax.yaxis.set_ticks([0,1])
    ax.set_yticklabels([0,1])
    ax.set_xticklabels([cols[-2], cols[-1]])

    plt.subplots_adjust(wspace=0)
    plt.legend([plt.Line2D((0,1), (0,0), color=colors[cat]) for cat in df['success'].cat.categories],
               df['success'].cat.categories,
               bbox_to_anchor=(1.2, 1), loc=2, borderaxespad=0.)
    plt.show() 

def plot_mean(df, title, axes):
    avg_dist = df.groupby('distance').mean()
    avg_width = df.groupby('width').mean()
    avg_beta = df.groupby('beta').mean()
    avg_length = df.groupby('length').mean()

    axes[0][0].plot(avg_dist.index, avg_dist['success'], label=title)
    axes[0][0].set_xlabel('Distance (m)')
    axes[0][0].set_ylabel('Average Success %')
    

    axes[0][1].plot(avg_width.index, avg_width['success'], label=title)
    axes[0][1].set_xlabel('Width (m)')

    axes[1][0].plot(avg_beta.index, avg_beta['success'], label=title)
    axes[1][0].set_xlabel('Beta (degrees)')
    axes[1][0].set_ylabel('Average Success %')


    axes[1][1].plot(avg_length.index, avg_length['success'], label=title)
    axes[1][1].set_xlabel('Length (m)')
    plt.legend()
    


def print_stats(df):
    df = df.groupby('success').mean()
    print(df)
    # print(f'0: {df['success'].iloc('0')}')

def parallel(df):
    FAIL = 0
    PASS = 1
    # df = df.groupby('success').agg(['mean', 'std'])
    df = df.groupby('success')
    df_m = df.mean()
    # print(df_m)
    # print(df.quantile(q=0.5))
    df_p = df_m.apply(lambda x: x+df.std()[x.name])
    df_2 = df_m.apply(lambda x: x+df.std()[x.name]/2)
    df_3 = df_m.apply(lambda x: x+df.std()[x.name]/3)
    print(df.std())

    # print(df_2)
    df_m = pandas.concat([df_m, df_p, df_2, df_3], ignore_index=False)
    print(df_m)
    # print(df_n)
    # x = [1, 2] # Four parameters
    # fig, axes = plt.subplots(1, 3, sharey=False)
    # y1 = [df['width'][FAIL], df['length'][FAIL]]
    # y2 = [df['width'][PASS], df['length'][PASS]] 
    # y3 = [df['length'][FAIL], df['beta'][FAIL]]
    # y4 = [df['length'][PASS], df['beta'][PASS]]
    # y5 = [df['beta'][FAIL], df['distance'][FAIL]]
    # y6 = [df['beta'][PASS], df['beta'][PASS]] 
    # axes[0].plot(x, y1, 'g-', x, y2, 'r-')
    # axes[1].plot(x, y3, 'g-', x, y4, 'r-')
    # axes[2].plot(x, y5, 'g-', x, y6, 'r-')


    # plt.subplots_adjust(wspace=0)
    # plt.show()
    # # # print(df['width'][FAIL])
    # print(df['success'])
    outcome_list = [val for pair in zip([FAIL], [PASS]) for _ in range( len(df_m) //2) for val in pair]
    width_std = [round(i, 2) for i in df['width'].std().values]
    length_std = [round(i, 2) for i in df['length'].std().values]
    beta_std = [round(i, 2) for i in df['beta'].std().values]
    distance_std = [round(i, 2) for i in df['distance'].std().values]


    fig = go.Figure(data=
                    go.Parcoords(
                        line=dict(color=outcome_list,
                                  colorscale=['red', 'blue']
                                  ),
                        dimensions = list([
                            dict(range = [0,4],
                                
                                label=f'Width: σ=[✔️:{width_std[1]}, ✘: {width_std[0]}]',
                                values=df_m['width']),
                            
                            dict(range=[0,20],
                                label=f'Length: σ=[✔️:{length_std[1]}, ✘: {length_std[0]}]',
                                values=df_m['length']),
                            
                            dict(range=[0,70],
                                 label=f'Beta: σ=[✔️:{beta_std[1]}, ✘: {beta_std[0]}]', 
                                 values=df_m['beta']),
                            
                            dict(range=[0,4],
                                 label=f'Distance: σ=[✔️:{distance_std[1]}, ✘: {distance_std[0]}]', 
                                 values=df_m['distance']),   
                        ])
                    ))
    fig.update_traces(legendgrouptitle=dict(
        font=dict(
            family='Times New Roman',
            size=12
        ),
        text='Test'
    ), selector=dict(type='parcoords'))

    fig.update_traces(line_colorbar=dict(
        len=1,
        showticklabels=True,
        dtick=0,
        nticks=2,
        labelalias={FAIL: 'Fail', PASS: 'Success'}
    ), selector=dict(type='parcoords'))

    fig.update_traces(labelfont=dict(
        family='Times New Roman',
        size=12
    ), selector='parcoords')

    fig.update_traces(tickfont=dict(
        family='Times New Roman',
        size=12
    ), selector='parcoords')

    fig.update_traces(rangefont=dict(
        family='Times New Roman',
        size=12
    ), selector='parcoords')
    

    fig.write_image('CLRerNetpc.pdf')
    fig.show()

if __name__ == '__main__':
    # print('here') default. Drag the lines along the axes to filter regions.


    # clean_data()
    # results_to_csv('./Shadow_Attack/masterscript/data/results', './Shadow_Attack/masterscript/data/')
    # results_to_csv('../data/results', '../data/')
    # plot_values()
    df_c, df_h, df_t = combine_dfs()
    # fig, axes = plt.subplots(2, 2,figsize=(10,10), sharey=False)

    plt.rcParams["font.family"] = "Times New Roman"
    plt.rcParams['font.size'] = 12

    # plot_mean(df_c, 'CLRerNet', axes)
    # plot_mean(df_h, 'HybridNets', axes)
    # plot_mean(df_t, 'TwinLiteNet', axes)
    # plt.subplots_adjust(left=0.065, bottom=0.055, right=0.984, top=0.968, wspace=0.197, hspace=0.2)
    # plt.savefig('Systevalmeans.pdf', format='pdf')
    # plt.legend()
    # plt.show()
    # print_stats(df_h)
    parallel(df_c)
    # parallel_plt(df_c)



