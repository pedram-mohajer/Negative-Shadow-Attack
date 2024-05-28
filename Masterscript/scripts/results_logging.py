import os
from openpyxl import Workbook

def create_workbook() -> Workbook:
    return Workbook()

def write_results(workbook: Workbook, sheet_name: str, data: dict, hide_fails=True) -> None:
    """ Writes results from data in to workbook via sheet_name

    Args:
        workbook: Workbook object to write data to
        sheet_name: Name of sheet for data to be written to
        data: data to write to sheet
        hide_fails: Whether or not to collapse rows with unsuccessful attacks
    """
    ROW_TITLES = ['Folder Name', 'Image', 'Attack Type', 'Success?', 'Success Type']
    sheet = workbook.create_sheet(title=sheet_name)
    sheet.append(ROW_TITLES)
    for folder in data:
        for image in data[folder]:
            for attack_type in data[folder][image]:
                success, success_type = data[folder][image][attack_type]
                row_data = [folder, image, attack_type, '1' if success is True else '0', success_type]
                sheet.append(row_data)
                
                if hide_fails and not success:
                    sheet.row_dimensions[sheet.max_row].hidden = True

def save_workbook(workbook: Workbook, file_path:  os.path) -> None:
    """ Saves workbook to specified file path
    
    Args:
        workbook: Workbook object to save
        file_path: file path to save workbook to
    """
    workbook.remove(workbook['Sheet'])
    workbook.save(file_path)
